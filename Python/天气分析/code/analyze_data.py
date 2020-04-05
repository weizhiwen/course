from openpyxl import load_workbook
from time import strptime
from matplotlib import pyplot

# 3、分析数据
# 读取数据
# 加载Excel文件
excel_url = '/tmp/南京历史天气预报.xlsx'
workbook = load_workbook(excel_url)
# 循环处理Excel的每个Sheet工作表
high_list = []
low_list = []
date_days_list = []
sheetnames = workbook.sheetnames
for sheetname in ['2019年3月份']:
	worksheet = workbook[sheetname]
	# 从第二行开始读取数据内容
	for row in worksheet.iter_rows(2, values_only=True):
		date = row[0]
		date_days = strptime(date, "%Y年%m月%d日").tm_yday
		date_days_list.append(date_days)
		temperature = row[2]
		high, low = temperature.replace('℃', '').split('/')
		# print(high, low)
		high_list.append(int(high))
		low_list.append(int(low))
	break
print('天数: ', date_days_list)
print('最高气温: ', high_list)
print('最低气温: ', low_list)

# 分析数据
# 画折线图
pyplot.plot(date_days_list, high_list, 'r-', label='Maximum temperature')
pyplot.plot(date_days_list, low_list, 'b-', label='Lowest temperature')
# 设置x轴刻度
x_list = []
for x in range(min(date_days_list), len(date_days_list) + 1):
    if (x - 1) % 10 == 0:
        x_list.append(x)
# 设置y轴刻度
pyplot.xticks(x_list)
y_list = []
for y in range(min(low_list), max(high_list) + 1):
    if (y - 1) % 2 == 0:
        y_list.append(y)
pyplot.yticks(y_list)
# 设置x、y轴描述
pyplot.xlabel("days")
pyplot.ylabel("˚C")
# 设置图标标题
pyplot.title("Shanghai weather in 2020")
# pyplot.grid(True)
# 绘制
pyplot.legend()
# 保存图片
pyplot.savefig("/tmp/{}.png".format(excel_url.split("/")[-1].replace('.xlsx', '')), format='png', transparent=False, dpi=300, pad_inches = 0)
